"""
Excel (3 sheets) and chart generation for BankNifty Short Strangle Backtest.

Sheet 1: Guide — strategy overview, parameter table, column definitions
Sheet 2: Tradesheet — formatted trade log with conditional coloring
Sheet 3: Statistics — all metrics, tables, and embedded charts

Charts: equity_curve.png and drawdown.png (300 DPI, dark professional theme)
"""

import logging
import os

import matplotlib
matplotlib.use("Agg")  # Non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

from config import (
    BASE_NAV,
    INITIAL_CAPITAL,
    EQUITY_CURVE_PNG,
    DRAWDOWN_PNG,
    EXCEL_FILENAME,
    OUTPUT_DIR,
    TARGET_PREMIUM,
    STOP_LOSS_PCT,
    LOT_SIZE,
    NUM_LOTS,
    QUANTITY,
    ENTRY_TIME,
    EXIT_TIME,
    EXPIRY_WEEKDAY,
)

logger = logging.getLogger(__name__)

# ─── Style constants ────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, italic=True, color="FFFFFF", size=11)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(color="276221")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006")
ORANGE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
SECTION_FONT = Font(bold=True, size=12, color="1F4E79")
BOLD_FONT = Font(bold=True, size=11)


def generate_charts(trades: pd.DataFrame, analytics: dict) -> None:
    """
    Generate equity curve and drawdown charts with dark professional theme.

    Args:
        trades: Trades DataFrame with EntryDate, NAV.
        analytics: Analytics dict with drawdown_series, nav_series, rolling_peak.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    dates = pd.to_datetime(trades["EntryDate"]).values
    nav_series = analytics["nav_series"]
    rolling_peak = analytics["rolling_peak"]
    drawdown_series = analytics["drawdown_series"]

    # ── Chart 1: Equity Curve ────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(14, 6))
    ax.plot(dates, nav_series, color="#00C8FF", linewidth=1.5, label="Strategy NAV")
    ax.fill_between(
        dates, rolling_peak, nav_series,
        where=(nav_series < rolling_peak),
        alpha=0.3, color="red", label="Drawdown periods",
    )
    ax.axhline(y=BASE_NAV, color="white", linestyle="--", linewidth=0.8, alpha=0.5, label=f"Base NAV ({BASE_NAV})")

    # Annotate max drawdown point
    max_dd_idx = np.argmin(drawdown_series)
    ax.annotate(
        f"Max DD: {drawdown_series[max_dd_idx]:.2f}%",
        xy=(dates[max_dd_idx], nav_series[max_dd_idx]),
        xytext=(30, -30), textcoords="offset points",
        fontsize=9, color="yellow",
        arrowprops=dict(arrowstyle="->", color="yellow"),
    )

    ax.set_facecolor("#1a1a2e")
    fig.patch.set_facecolor("#1a1a2e")
    ax.set_title("BankNifty Short Strangle — Equity Curve (NAV)", color="white", fontsize=14, pad=15)
    ax.set_xlabel("Date", color="white")
    ax.set_ylabel("NAV", color="white")
    ax.tick_params(colors="white")
    ax.legend(loc="upper left", facecolor="#1a1a2e", edgecolor="white", labelcolor="white")
    ax.grid(True, alpha=0.2, color="white")
    ax.spines["bottom"].set_color("white")
    ax.spines["left"].set_color("white")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    fig.tight_layout()
    fig.savefig(EQUITY_CURVE_PNG, dpi=300, facecolor="#1a1a2e")
    plt.close(fig)
    logger.info("Equity curve saved to %s", EQUITY_CURVE_PNG)

    # ── Chart 2: Drawdown ────────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(14, 4))
    ax.fill_between(dates, drawdown_series, 0, color="red", alpha=0.6)
    ax.plot(dates, drawdown_series, color="#FF4444", linewidth=1)

    max_dd_pct = drawdown_series[max_dd_idx]
    ax.axhline(y=max_dd_pct, color="yellow", linestyle="--", linewidth=1, label=f"Max DD: {max_dd_pct:.2f}%")
    ax.annotate(
        f"{max_dd_pct:.2f}%",
        xy=(dates[max_dd_idx], max_dd_pct),
        xytext=(30, -20), textcoords="offset points",
        fontsize=9, color="yellow",
        arrowprops=dict(arrowstyle="->", color="yellow"),
    )

    ax.set_facecolor("#1a1a2e")
    fig.patch.set_facecolor("#1a1a2e")
    ax.set_title("BankNifty Short Strangle — Drawdown", color="white", fontsize=14, pad=15)
    ax.set_xlabel("Date", color="white")
    ax.set_ylabel("Drawdown %", color="white")
    ax.tick_params(colors="white")
    ax.legend(loc="lower left", facecolor="#1a1a2e", edgecolor="white", labelcolor="white")
    ax.grid(True, alpha=0.2, color="white")
    ax.spines["bottom"].set_color("white")
    ax.spines["left"].set_color("white")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    fig.tight_layout()
    fig.savefig(DRAWDOWN_PNG, dpi=300, facecolor="#1a1a2e")
    plt.close(fig)
    logger.info("Drawdown chart saved to %s", DRAWDOWN_PNG)


def _write_guide_sheet(ws) -> None:
    """Write the Guide sheet with strategy documentation."""
    ws.title = "Guide"
    ws.sheet_properties.tabColor = "1F4E79"

    rows = [
        ("BankNifty Short Strangle Backtester", SECTION_FONT),
        ("", None),
        ("STRATEGY OVERVIEW", SECTION_FONT),
        ("This backtester implements a daily short strangle strategy on BankNifty index options.", None),
        ("At 09:20 each trading day, we sell (short) one CE and one PE option with premium closest to Rs. 50.", None),
        ("Each leg has an independent 50% stop loss. Normal exit at 15:20.", None),
        ("Fixed lot size of 15 units per leg. No compounding.", None),
        ("", None),
        ("PARAMETERS", SECTION_FONT),
    ]

    for i, (text, font) in enumerate(rows, 1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font

    # Parameter table
    params = [
        ("Parameter", "Value", "Description"),
        ("Entry Time", ENTRY_TIME, "1-min bar close time for entry"),
        ("Exit Time", EXIT_TIME, "1-min bar close time for normal exit"),
        ("Target Premium", f"Rs. {TARGET_PREMIUM}", "Strike selection target"),
        ("Stop Loss", f"{STOP_LOSS_PCT*100:.0f}%", "Above entry price per leg"),
        ("Lot Size", str(LOT_SIZE), "BankNifty lot size"),
        ("Num Lots", str(NUM_LOTS), "Fixed lots per day per leg"),
        ("Quantity", str(QUANTITY), "LotSize × NumLots"),
        ("Initial Capital", f"Rs. {INITIAL_CAPITAL:,.0f}", "Starting capital"),
        ("Base NAV", str(BASE_NAV), "Equity curve base index"),
        ("Expiry Weekday", "Wednesday", "Week 1 expiry day"),
    ]

    start_row = 10
    for i, (p, v, d) in enumerate(params):
        r = start_row + i
        ws.cell(row=r, column=1, value=p).font = BOLD_FONT if i == 0 else Font()
        ws.cell(row=r, column=2, value=v).font = BOLD_FONT if i == 0 else Font()
        ws.cell(row=r, column=3, value=d).font = BOLD_FONT if i == 0 else Font()
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if i == 0:
                ws.cell(row=r, column=c).fill = HEADER_FILL
                ws.cell(row=r, column=c).font = HEADER_FONT

    # Column definitions
    col_start = start_row + len(params) + 2
    ws.cell(row=col_start, column=1, value="TRADESHEET COLUMN DEFINITIONS").font = SECTION_FONT

    col_defs = [
        ("Column", "Description"),
        ("Ticker", "Full option ticker e.g. BANKNIFTY43800CE"),
        ("Option_Type", "CE or PE"),
        ("EntryDate", "Date of trade entry"),
        ("EntryTime", "Always 09:20:59"),
        ("EntryPrice", "09:20 close of selected option"),
        ("ExitDate", "Date of exit (same day — intraday)"),
        ("ExitTime", "15:20:59 or time of SL hit"),
        ("ExitPrice", "15:20 close OR EntryPrice × 1.50 (SL)"),
        ("ExitReason", "Regular or StopLoss"),
        ("Banknifty_Close", "Spot close at 09:20 entry bar"),
        ("LotSize", "15 (BankNifty lot size)"),
        ("Quantity", "15 (LotSize × NumLots)"),
        ("EntryValue", "EntryPrice × Quantity"),
        ("ExitValue", "ExitPrice × Quantity"),
        ("GrossPnL", "(EntryPrice - ExitPrice) × Quantity"),
        ("CumulativePnL", "Running cumulative sum of GrossPnL"),
        ("AvailableCapital", "InitialCapital + CumulativePnL"),
        ("IsExpiryDay", "True if Wednesday in first week of month"),
        ("Strike", "Numeric strike price extracted from ticker"),
        ("NAV", "Equity curve value (base 100)"),
    ]

    for i, (col_name, desc) in enumerate(col_defs):
        r = col_start + 1 + i
        ws.cell(row=r, column=1, value=col_name).font = BOLD_FONT if i == 0 else Font()
        ws.cell(row=r, column=2, value=desc).font = BOLD_FONT if i == 0 else Font()
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if i == 0:
                ws.cell(row=r, column=c).fill = HEADER_FILL
                ws.cell(row=r, column=c).font = HEADER_FONT

    # SL explanation
    sl_start = col_start + len(col_defs) + 3
    ws.cell(row=sl_start, column=1, value="STOP LOSS MECHANISM").font = SECTION_FONT
    sl_lines = [
        "SL Price = EntryPrice × 1.50 (50% above entry, since we are SHORT)",
        "After entry, scan every 1-min HIGH bar from 09:21 to 15:20",
        "If any bar's HIGH >= SL Price: leg is stopped out at SL Price exactly",
        "We use HIGH (not Close) because intrabar price can touch SL even if bar closes below",
        "Each leg is independent — one hitting SL does not affect the other",
    ]
    for i, line in enumerate(sl_lines):
        ws.cell(row=sl_start + 1 + i, column=1, value=line)

    # Auto-fit column widths
    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 40


def _write_tradesheet(ws, trades: pd.DataFrame) -> None:
    """Write the Tradesheet with formatting and conditional coloring."""
    ws.title = "Tradesheet"
    ws.sheet_properties.tabColor = "27AE60"

    # Select display columns
    display_cols = [
        "Ticker", "Option_Type", "EntryDate", "EntryTime", "EntryPrice",
        "ExitDate", "ExitTime", "ExitPrice", "ExitReason",
        "Banknifty_Close", "LotSize", "Quantity",
        "EntryValue", "ExitValue", "GrossPnL",
        "CumulativePnL", "AvailableCapital", "IsExpiryDay", "Strike",
    ]
    display_cols = [c for c in display_cols if c in trades.columns]

    # Header row
    for j, col_name in enumerate(display_cols, 1):
        cell = ws.cell(row=1, column=j, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Freeze header
    ws.freeze_panes = "A2"

    # Data rows
    for i, (_, row) in enumerate(trades.iterrows(), 2):
        pnl = row.get("GrossPnL", 0)
        exit_reason = row.get("ExitReason", "Regular")

        for j, col_name in enumerate(display_cols, 1):
            val = row[col_name]
            # Convert numpy types
            if hasattr(val, "item"):
                val = val.item()
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

            # Number formatting
            if col_name in ("EntryPrice", "ExitPrice", "Banknifty_Close"):
                cell.number_format = "#,##0.00"
            elif col_name in ("GrossPnL", "CumulativePnL"):
                cell.number_format = "+#,##0.00;-#,##0.00"
            elif col_name in ("EntryValue", "ExitValue", "AvailableCapital"):
                cell.number_format = "#,##0.00"

        # Row coloring
        if exit_reason == "StopLoss":
            fill = ORANGE_FILL
        elif pnl > 0:
            fill = GREEN_FILL
            font = GREEN_FONT
        else:
            fill = RED_FILL
            font = RED_FONT

        for j in range(1, len(display_cols) + 1):
            ws.cell(row=i, column=j).fill = fill
            if exit_reason != "StopLoss":
                ws.cell(row=i, column=j).font = font if pnl > 0 else RED_FONT

        # Alternating gray for non-colored rows (overridden by PnL colors above,
        # but we keep the pattern for visual reference on SL rows)

    # Bold totals row
    total_row = len(trades) + 2
    ws.cell(row=total_row, column=1, value="TOTALS").font = Font(bold=True, italic=True, size=11)
    # Find GrossPnL column index
    if "GrossPnL" in display_cols:
        pnl_col = display_cols.index("GrossPnL") + 1
        total_pnl = trades["GrossPnL"].sum()
        cell = ws.cell(row=total_row, column=pnl_col, value=total_pnl)
        cell.font = Font(bold=True, size=11)
        cell.number_format = "+#,##0.00;-#,##0.00"
        cell.border = THIN_BORDER

    # Auto-fit column widths
    for j, col_name in enumerate(display_cols, 1):
        max_len = max(len(str(col_name)), 12)
        ws.column_dimensions[get_column_letter(j)].width = max_len + 4


def _write_statistics_sheet(ws, trades: pd.DataFrame, analytics: dict) -> None:
    """Write the Statistics sheet with all metrics, tables, and embedded charts."""
    ws.title = "Statistics"
    ws.sheet_properties.tabColor = "E74C3C"

    row = 1

    # ── Section 1: Return Metrics ────────────────────────────────────────
    ws.cell(row=row, column=1, value="RETURN METRICS").font = SECTION_FONT
    row += 1
    metrics = [
        ("CAGR", f"{analytics['cagr']*100:.2f}%"),
        ("Total Return", f"{analytics['total_return_pct']:.2f}%"),
        ("Final Capital", f"Rs. {analytics['final_capital']:,.2f}"),
        ("Avg Daily PnL", f"Rs. {analytics['mean_daily_pnl']:,.2f}"),
        ("Avg Trade PnL", f"Rs. {analytics['avg_trade_pnl']:,.2f}"),
    ]
    for label, val in metrics:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # ── Section 2: Risk Metrics ──────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="RISK METRICS").font = SECTION_FONT
    row += 1
    risk = [
        ("Max Drawdown", f"{analytics['max_drawdown_pct']:.2f}%"),
        ("Max DD Date", str(analytics["max_drawdown_date"])),
        ("Sharpe Ratio", f"{analytics['sharpe_ratio']:.4f}"),
        ("Sortino Ratio", f"{analytics['sortino_ratio']:.4f}"),
        ("Calmar Ratio", f"{analytics['calmar_ratio']:.4f}"),
    ]
    for label, val in risk:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # ── Section 3: Win/Loss Table ────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="WIN / LOSS ANALYSIS").font = SECTION_FONT
    row += 1
    headers = ["Group", "Winners", "Losers", "Total", "Win %", "Loss %"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row += 1

    wl = analytics["win_loss"]
    for group in ["CE", "PE", "Combined"]:
        d = wl[group]
        vals = [group, d["winners"], d["losers"], d["total"], f"{d['win_pct']:.1f}%", f"{d['loss_pct']:.1f}%"]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=j, value=v)
            cell.border = THIN_BORDER
        row += 1

    # ── Section 4: Avg % P&L by Expiry ──────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="AVG % P&L BY OPTION TYPE & EXPIRY").font = SECTION_FONT
    row += 1
    headers = ["Category", "Avg % PnL"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row += 1

    for key, val in analytics["avg_pnl_expiry"].items():
        ws.cell(row=row, column=1, value=key).border = THIN_BORDER
        cell = ws.cell(row=row, column=2, value=f"{val:.2f}%")
        cell.border = THIN_BORDER
        cell.fill = GREEN_FILL if val > 0 else RED_FILL
        row += 1

    # ── Section 5: Monthly P&L ───────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="MONTHLY P&L").font = SECTION_FONT
    row += 1
    monthly = analytics["monthly_pnl"]
    if not monthly.empty:
        m_headers = ["Month", "Start NAV", "End NAV", "% P&L", "Abs P&L (Rs.)"]
        for j, h in enumerate(m_headers, 1):
            cell = ws.cell(row=row, column=j, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        row += 1

        for _, mrow in monthly.iterrows():
            ws.cell(row=row, column=1, value=mrow["Month"]).border = THIN_BORDER
            ws.cell(row=row, column=2, value=mrow["StartNAV"]).border = THIN_BORDER
            ws.cell(row=row, column=3, value=mrow["EndNAV"]).border = THIN_BORDER

            pct_cell = ws.cell(row=row, column=4, value=mrow["PctPnL"])
            pct_cell.border = THIN_BORDER
            pct_cell.number_format = "+#,##0.00;-#,##0.00"
            pct_cell.fill = GREEN_FILL if mrow["PctPnL"] > 0 else RED_FILL

            abs_cell = ws.cell(row=row, column=5, value=mrow["AbsPnL"])
            abs_cell.border = THIN_BORDER
            abs_cell.number_format = "+#,##0.00;-#,##0.00"
            row += 1

    # ── Section 6: Additional Stats ──────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="ADDITIONAL STATISTICS").font = SECTION_FONT
    row += 1
    extra = [
        ("Total SL Hits", analytics["total_sl_hits"]),
        ("CE SL Hits", analytics["ce_sl_hits"]),
        ("PE SL Hits", analytics["pe_sl_hits"]),
        ("Avg Premium Collected", f"Rs. {analytics['avg_premium_collected']:.2f}"),
        ("Max Single Trade Profit", f"Rs. {analytics['max_single_profit']:.2f}"),
        ("Max Single Trade Loss", f"Rs. {analytics['max_single_loss']:.2f}"),
        ("Total Trading Days", analytics["total_trading_days"]),
        ("Total Legs Traded", analytics["total_legs_traded"]),
    ]
    for label, val in extra:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # ── Section 7 & 8: Embedded Charts ───────────────────────────────────
    row += 2
    ws.cell(row=row, column=1, value="EQUITY CURVE").font = SECTION_FONT
    row += 1
    if os.path.exists(EQUITY_CURVE_PNG):
        img = XLImage(EQUITY_CURVE_PNG)
        img.width = 900
        img.height = 400
        ws.add_image(img, f"A{row}")
        row += 22  # leave space

    ws.cell(row=row, column=1, value="DRAWDOWN CHART").font = SECTION_FONT
    row += 1
    if os.path.exists(DRAWDOWN_PNG):
        img = XLImage(DRAWDOWN_PNG)
        img.width = 900
        img.height = 280
        ws.add_image(img, f"A{row}")

    # Auto-fit columns
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20


def generate_excel_report(trades: pd.DataFrame, analytics: dict) -> None:
    """
    Generate the full Excel report with 3 sheets.

    Args:
        trades: Final trades DataFrame.
        analytics: Analytics dict from performance_analytics.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()

    # Sheet 1: Guide
    ws_guide = wb.active
    _write_guide_sheet(ws_guide)

    # Sheet 2: Tradesheet
    ws_trades = wb.create_sheet()
    _write_tradesheet(ws_trades, trades)

    # Sheet 3: Statistics
    ws_stats = wb.create_sheet()
    _write_statistics_sheet(ws_stats, trades, analytics)

    wb.save(EXCEL_FILENAME)
    logger.info("Excel report saved to %s", EXCEL_FILENAME)
